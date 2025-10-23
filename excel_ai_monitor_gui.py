"""Excel AI Monitor GUI.

This module provides a PySimpleGUI-based desktop application for monitoring
multiple Excel workbooks.  Each monitored workbook is periodically sampled and
sent to a Large Language Model (LLM) via the official :mod:`openai` SDK so that
the model can decide which rows are "meaningful".  The UI shows the latest
analysis summary, surfaces smart alerts, renders a live numeric trend graph,
and allows manual re-analysis on demand.

The application is intentionally self-contained so that it can run without an
actual LLM endpoint during development.  If no API is configured the monitor
falls back to a deterministic heuristic analysis which computes basic
statistics for numeric columns.  A lightweight HTTP compatibility mode is kept
for custom gateways, and a Teams-style notifier (``# call via team``) marks
events that should be routed to external collaboration tooling.

Requirements
------------
* PySimpleGUI
* openpyxl
* openai (optional but recommended)
* matplotlib (optional, for visualisations)
* requests (only for custom HTTP gateways)
"""

from __future__ import annotations

import json
import re
import threading
import time
from collections import deque
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from queue import Queue, Empty
from statistics import mean
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import PySimpleGUI as sg
try:  # Optional dependency: OpenAI SDK for LLM access.
    import openai
except ImportError:  # pragma: no cover - optional dependency
    openai = None
import requests
from openpyxl import load_workbook

try:  # Optional dependency: Matplotlib for visualization.
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure

    _MATPLOTLIB_AVAILABLE = True
except Exception:  # pragma: no cover - optional dependency
    FigureCanvasTkAgg = None  # type: ignore
    Figure = None  # type: ignore
    _MATPLOTLIB_AVAILABLE = False


# ---------------------------------------------------------------------------
# LLM client helpers
# ---------------------------------------------------------------------------


@dataclass
class LLMResult:
    """Simple container for responses coming back from the LLM client."""

    summary: str
    meaningful_rows: List[dict] = field(default_factory=list)
    raw_response: Optional[dict] = None


class LLMAnalyzer:
    """Send workbook samples to an LLM using the OpenAI SDK when available."""

    def __init__(
        self,
        api_url: str = "",
        api_key: str = "",
        timeout: int = 60,
        model: str = "gpt-4o-mini",
    ):
        self.api_url = api_url.strip()
        self.api_key = api_key.strip()
        self.timeout = max(1, int(timeout))
        self.model = model.strip() or "gpt-4o-mini"
        self._client = None

    def analyze_dataset(
        self,
        file_path: str,
        sheet_name: str,
        rows: Iterable[dict],
    ) -> LLMResult:
        """Analyze the provided ``rows`` and return an :class:`LLMResult`.

        The analyzer prefers to use the :mod:`openai` SDK so that prompts can
        take advantage of the user's configured Large Language Model.  When no
        valid client configuration is present the method falls back to a
        deterministic heuristic analysis which keeps the rest of the UI
        responsive.
        """

        rows_list = list(rows)
        if not rows_list:
            return LLMResult(summary="No data available in the sheet.")

        # Try the OpenAI SDK first when available/configured.
        if openai is not None and (self.api_key or getattr(openai, "api_key", None)):
            result = self._analyze_with_openai(file_path, sheet_name, rows_list)
            if result:
                return result

        # As a compatibility path, honour explicit HTTP endpoints (e.g. custom
        # gateways or mock servers) before falling back to heuristics.
        if self.api_url:
            http_result = self._analyze_via_http(file_path, sheet_name, rows_list)
            if http_result:
                return http_result

        return self._fallback_analysis(file_path, sheet_name, rows_list)

    # ------------------------------------------------------------------
    # OpenAI helpers
    # ------------------------------------------------------------------

    def _ensure_openai_client(self):
        if openai is None:
            return None
        if self._client is not None:
            return self._client

        try:
            if hasattr(openai, "OpenAI"):
                kwargs = {}
                if self.api_key:
                    kwargs["api_key"] = self.api_key
                if self.api_url:
                    kwargs["base_url"] = self.api_url
                self._client = openai.OpenAI(**kwargs)
            else:  # Legacy SDK (<1.0.0)
                if self.api_key:
                    openai.api_key = self.api_key
                if self.api_url:
                    setattr(openai, "api_base", self.api_url)
                self._client = openai
        except Exception:  # pragma: no cover - defensive client init
            self._client = None
        return self._client

    def _system_prompt(self) -> str:
        return (
            "You are an electrical systems data analyst. Interpret Excel rows "
            "that contain power, voltage, and process telemetry. Identify "
            "noteworthy operating states (faults, recoveries, important "
            "progress milestones) and explain why they matter. Always reply "
            "with strict JSON containing 'summary' (string) and "
            "'meaningful_rows' (list of objects with at least 'row_number' and "
            "'explanation')."
        )

    def _analyze_with_openai(
        self, file_path: str, sheet_name: str, rows: List[dict]
    ) -> Optional[LLMResult]:
        client = self._ensure_openai_client()
        if client is None:
            return None

        prompt = self._build_prompt(file_path, sheet_name, rows)

        try:
            if hasattr(client, "responses"):
                response = client.responses.create(
                    model=self.model,
                    input=[
                        {"role": "system", "content": self._system_prompt()},
                        {"role": "user", "content": prompt},
                    ],
                    max_output_tokens=800,
                    timeout=self.timeout,
                )
                raw_response = response.model_dump() if hasattr(response, "model_dump") else {}
                content = getattr(response, "output_text", "") or raw_response.get("output_text", "")
            elif hasattr(client, "chat") and hasattr(client.chat, "completions"):
                response = client.chat.completions.create(
                    model=self.model,
                    messages=[
                        {"role": "system", "content": self._system_prompt()},
                        {"role": "user", "content": prompt},
                    ],
                    temperature=0.1,
                    timeout=self.timeout,
                )
                raw_response = response.dict() if hasattr(response, "dict") else response.model_dump()
                content = response.choices[0].message.content  # type: ignore[attr-defined]
            else:  # Legacy client
                response = client.ChatCompletion.create(
                    model=self.model,
                    messages=[
                        {"role": "system", "content": self._system_prompt()},
                        {"role": "user", "content": prompt},
                    ],
                    temperature=0.1,
                    request_timeout=self.timeout,
                )
                raw_response = response
                content = response["choices"][0]["message"]["content"]
        except Exception as exc:  # pragma: no cover - network
            fallback = self._fallback_analysis(file_path, sheet_name, rows)
            summary = (
                f"LLM request failed ({exc}). Falling back to heuristic analysis.\n\n"
                f"{fallback.summary}"
            )
            return LLMResult(
                summary=summary,
                meaningful_rows=fallback.meaningful_rows,
                raw_response={"error": str(exc)},
            )

        parsed = self._parse_llm_payload(content)
        summary = (
            parsed.get("summary")
            or parsed.get("message")
            or parsed.get("result")
            or "Received response without a summary."
        )
        meaningful = parsed.get("meaningful_rows") or parsed.get("rows") or []
        if isinstance(meaningful, dict):
            meaningful = [meaningful]

        return LLMResult(
            summary=str(summary),
            meaningful_rows=list(meaningful),
            raw_response={"openai": raw_response, "parsed": parsed},
        )

    def _parse_llm_payload(self, content: str) -> dict:
        if not content:
            return {}

        candidate = content.strip()
        try:
            return json.loads(candidate)
        except json.JSONDecodeError:
            match = re.search(r"\{.*\}", candidate, re.DOTALL)
            if match:
                try:
                    return json.loads(match.group(0))
                except json.JSONDecodeError:
                    pass
        return {"summary": candidate}

    # ------------------------------------------------------------------
    # HTTP compatibility helper
    # ------------------------------------------------------------------

    def _analyze_via_http(
        self, file_path: str, sheet_name: str, rows: List[dict]
    ) -> Optional[LLMResult]:
        headers = {"Content-Type": "application/json"}
        if self.api_key:
            headers["Authorization"] = f"Bearer {self.api_key}"

        prompt = self._build_prompt(file_path, sheet_name, rows)
        payload = {"prompt": prompt, "max_tokens": 512}

        try:
            response = requests.post(
                self.api_url,
                headers=headers,
                json=payload,
                timeout=self.timeout,
            )
            response.raise_for_status()
        except requests.RequestException:
            return None

        try:
            data = response.json()
        except ValueError:
            data = {"summary": response.text, "meaningful_rows": []}

        summary = (
            data.get("summary")
            or data.get("message")
            or data.get("result")
            or "Received response without a summary."
        )
        meaningful = data.get("meaningful_rows") or data.get("rows") or []
        if isinstance(meaningful, dict):
            meaningful = [meaningful]

        return LLMResult(summary=summary, meaningful_rows=list(meaningful), raw_response=data)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _build_prompt(self, file_path: str, sheet_name: str, rows: List[dict]) -> str:
        safe_name = sheet_name or "(first sheet)"
        json_rows = json.dumps(rows, indent=2, default=str)
        prompt = (
            "You are an analyst that identifies meaningful patterns in Excel data. "
            "Determine the key insights and list the rows that best illustrate them."
            "\n\n"
            f"Workbook: {file_path}\n"
            f"Sheet: {safe_name}\n"
            "Data sample (JSON, includes `row_number` and `values`):\n"
            f"{json_rows}\n\n"
            "Respond with JSON containing `summary` (human readable text) and "
            "`meaningful_rows` (list of objects describing the most important rows)."
        )
        return prompt

    def _fallback_analysis(
        self, file_path: str, sheet_name: str, rows: List[dict]
    ) -> LLMResult:
        """Provide a deterministic heuristic analysis when no API is configured."""

        numeric_columns: Dict[str, List[float]] = {}
        for item in rows:
            values = item.get("values", {})
            for key, raw in values.items():
                try:
                    number = float(raw)
                except (TypeError, ValueError):
                    continue
                numeric_columns.setdefault(key, []).append(number)

        lines = [
            f"Workbook '{Path(file_path).name}' sheet '{sheet_name or 'first'}'"
            f" analysed locally (no LLM endpoint configured).",
            f"Rows sampled: {len(rows)}",
        ]

        meaningful: List[dict] = []

        if numeric_columns:
            lines.append("Numeric column averages:")
            for key, numbers in numeric_columns.items():
                lines.append(f"  • {key}: {mean(numbers):.2f}")

            # Choose rows with the highest sum across numeric fields as "meaningful"
            scored_rows = []
            for item in rows:
                values = item.get("values", {})
                score = 0.0
                for key in numeric_columns:
                    try:
                        score += float(values.get(key, 0))
                    except (TypeError, ValueError):
                        continue
                scored_rows.append((score, item))
            scored_rows.sort(key=lambda x: x[0], reverse=True)
            meaningful = [row for _, row in scored_rows[: min(5, len(scored_rows))]]
            if meaningful:
                lines.append(
                    "Top rows selected because they have the largest sum across "
                    "numeric columns."
                )
        else:
            lines.append(
                "No numeric columns detected; showing the first few rows as a "
                "representative sample."
            )
            meaningful = rows[: min(5, len(rows))]

        summary = "\n".join(lines)
        return LLMResult(summary=summary, meaningful_rows=meaningful)


# ---------------------------------------------------------------------------
# Excel monitoring worker threads
# ---------------------------------------------------------------------------


def _format_rows_for_display(data: object) -> str:
    """Return a pretty printed string representation of arbitrary JSON data."""

    return json.dumps(data, indent=2, default=str)


class ExcelMonitorWorker(threading.Thread):
    """Thread that watches one workbook/sheet and performs AI analysis."""

    def __init__(
        self,
        path: str,
        sheet: str = "",
        interval: int = 60,
        sample_rows: int = 25,
        api_url: str = "",
        api_key: str = "",
        model: str = "gpt-4o-mini",
        outq: Optional[Queue] = None,
    ):
        super().__init__(daemon=True)
        self.path = str(Path(path))
        self.sheet = sheet
        self.interval = max(1, int(interval))
        self.sample_rows = max(1, int(sample_rows))
        self.api_url = api_url
        self.api_key = api_key
        self.model = model
        self.outq = outq or Queue()

        self._stop = threading.Event()
        self._force = threading.Event()
        self._last_mtime: Optional[float] = None
        self._last_summary: str = "Analysis pending"
        self._last_meaningful: List[dict] = []
        self._last_preview: List[dict] = []
        self._last_signature: Optional[Tuple] = None
        self._analysis_count = 0

    # ------------------------------------------------------------------
    # Lifecycle helpers
    # ------------------------------------------------------------------

    def stop(self) -> None:
        self._stop.set()

    def trigger_manual_analysis(self) -> None:
        self._force.set()

    def update_settings(
        self,
        *,
        interval: Optional[int] = None,
        api_url: Optional[str] = None,
        api_key: Optional[str] = None,
        sample_rows: Optional[int] = None,
        model: Optional[str] = None,
    ) -> None:
        if interval is not None:
            self.interval = max(1, int(interval))
        if sample_rows is not None:
            self.sample_rows = max(1, int(sample_rows))
        if api_url is not None:
            self.api_url = api_url
        if api_key is not None:
            self.api_key = api_key
        if model is not None:
            self.model = model or self.model
        # Force a refresh so that new credentials take effect quickly.
        self._force.set()

    # ------------------------------------------------------------------
    # Worker loop
    # ------------------------------------------------------------------

    def run(self) -> None:  # pragma: no branch - thread loop is linear
        self._emit_status("Monitoring started")

        while not self._stop.is_set():
            run_analysis = self._force.is_set()

            try:
                path = Path(self.path)
                if not path.exists():
                    raise FileNotFoundError("Workbook not found")

                current_mtime = path.stat().st_mtime
                if self._last_mtime is None or current_mtime != self._last_mtime:
                    run_analysis = True

                if run_analysis:
                    preview_rows = self._load_preview()
                    analyzer = LLMAnalyzer(
                        self.api_url,
                        self.api_key,
                        model=self.model,
                    )
                    result = analyzer.analyze_dataset(self.path, self.sheet, preview_rows)
                    self._last_summary = result.summary
                    self._last_meaningful = list(result.meaningful_rows)
                    self._last_preview = list(preview_rows)
                    self._analysis_count += 1
                    signature = self._meaningful_signature(self._last_meaningful)
                    changed = signature != self._last_signature
                    self._last_signature = signature
                    status = "Analyzed" if changed else "Analyzed (stable)"
                    self._last_mtime = current_mtime
                else:
                    status = "No change"

                extra = {
                    "analysis_count": self._analysis_count,
                    "insight_changed": run_analysis and status == "Analyzed",
                    "meaningful_signature": self._last_signature,
                }
                self._emit_status(status, extra=extra)
            except Exception as exc:  # pragma: no cover - defensive
                self._emit_status("Error", error=str(exc))

            self._force.clear()

            # Sleep in 0.2s chunks so that stop/force reacts quickly
            for _ in range(self.interval * 5):
                if self._stop.is_set() or self._force.is_set():
                    break
                time.sleep(0.2)

        self._emit_status("Stopped")

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _emit_status(self, status: str, error: str = "", extra: Optional[dict] = None) -> None:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        message = {
            "file": self.path,
            "sheet": self.sheet if self.sheet else "(first sheet)",
            "status": status,
            "timestamp": timestamp,
            "summary": self._last_summary,
            "meaningful_rows": list(self._last_meaningful),
            "preview_rows": list(self._last_preview),
            "error": error,
            "analysis_count": self._analysis_count,
        }
        if extra:
            message.update(extra)
        self.outq.put(message)

    def _load_preview(self) -> List[dict]:
        wb = load_workbook(filename=self.path, read_only=True, data_only=True)
        try:
            if self.sheet:
                ws = wb[self.sheet]
            else:
                ws = wb[wb.sheetnames[0]]

            rows_iter = ws.iter_rows(values_only=True)
            try:
                headers = next(rows_iter)
            except StopIteration:
                return []

            headers = [h if h is not None else f"Column {i+1}" for i, h in enumerate(headers)]

            preview: List[dict] = []
            for idx, row in enumerate(rows_iter, start=2):
                row_data = {
                    header: row[col_idx] if col_idx < len(row) else None
                    for col_idx, header in enumerate(headers)
                }
                preview.append({"row_number": idx, "values": row_data})
                if len(preview) >= self.sample_rows:
                    break

            return preview
        finally:
            wb.close()

    def _meaningful_signature(self, rows: Sequence[dict]) -> Tuple:
        signature: List[Tuple] = []
        for row in rows:
            row_number = row.get("row_number")
            values = row.get("values", {})
            ordered = tuple(sorted(values.items())) if isinstance(values, dict) else tuple()
            signature.append((row_number, ordered))
        return tuple(signature)


# ---------------------------------------------------------------------------
# Monitor manager: keeps track of workers and surfaces queue updates
# ---------------------------------------------------------------------------


class ExcelMonitorManager:
    """Owns :class:`ExcelMonitorWorker` instances and exposes a common queue."""

    def __init__(self, interval: int = 60, sample_rows: int = 25, model: str = "gpt-4o-mini"):
        self.interval = max(1, int(interval))
        self.sample_rows = max(1, int(sample_rows))
        self.api_url = ""
        self.api_key = ""
        self.model = model
        self.outq: Queue = Queue()
        self._workers: Dict[tuple, ExcelMonitorWorker] = {}

    @staticmethod
    def _key(path: str, sheet: str) -> tuple:
        return (str(Path(path)), sheet or "")

    def update_global_settings(
        self,
        *,
        interval: Optional[int] = None,
        sample_rows: Optional[int] = None,
        api_url: Optional[str] = None,
        api_key: Optional[str] = None,
        model: Optional[str] = None,
    ) -> None:
        if interval is not None:
            self.interval = max(1, int(interval))
        if sample_rows is not None:
            self.sample_rows = max(1, int(sample_rows))
        if api_url is not None:
            self.api_url = api_url
        if api_key is not None:
            self.api_key = api_key
        if model is not None:
            self.model = model or self.model

        for worker in self._workers.values():
            worker.update_settings(
                interval=self.interval,
                sample_rows=self.sample_rows,
                api_url=self.api_url,
                api_key=self.api_key,
                model=self.model,
            )

    def add(self, path: str, sheet: str = "") -> bool:
        key = self._key(path, sheet)
        if key in self._workers:
            return False

        worker = ExcelMonitorWorker(
            path,
            sheet,
            interval=self.interval,
            sample_rows=self.sample_rows,
            api_url=self.api_url,
            api_key=self.api_key,
            model=self.model,
            outq=self.outq,
        )
        self._workers[key] = worker
        worker.start()
        return True

    def remove(self, key: tuple) -> bool:
        worker = self._workers.pop(key, None)
        if not worker:
            return False
        worker.stop()
        worker.join(timeout=2)
        return True

    def trigger_analysis(self, key: tuple) -> None:
        worker = self._workers.get(key)
        if worker:
            worker.trigger_manual_analysis()

    def stop_all(self) -> None:
        for worker in list(self._workers.values()):
            worker.stop()
        for worker in list(self._workers.values()):
            worker.join(timeout=2)
        self._workers.clear()

    def drain(self) -> List[dict]:
        messages: List[dict] = []
        try:
            while True:
                messages.append(self.outq.get_nowait())
        except Empty:
            pass
        return messages


# ---------------------------------------------------------------------------
# Alerting and visualisation helpers
# ---------------------------------------------------------------------------


def notify_team(event_type: str, message: str) -> None:
    """Placeholder Teams integration hook.

    The real environment is expected to replace this with a custom Teams
    connector.  The ``# call via team`` prefix is retained as requested so that
    downstream automation can grep for outbound alerts.
    """

    tag = event_type.upper() if event_type else "INFO"
    print(f"# call via team: [{tag}] {message}")


class AlertManager:
    """Decide when to push smart notifications about monitoring progress."""

    def __init__(self, history_limit: int = 200):
        self._history: deque[str] = deque(maxlen=history_limit)
        self._last_state: Dict[tuple, dict] = {}

    @property
    def history(self) -> List[str]:
        return list(self._history)

    def process(self, payload: dict) -> List[str]:
        key = (payload.get("file", ""), payload.get("sheet", ""))
        status = payload.get("status", "")
        summary = payload.get("summary", "")
        analysis_count = int(payload.get("analysis_count") or 0)
        changed = bool(payload.get("insight_changed"))
        error = payload.get("error", "")
        timestamp = payload.get("timestamp") or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rows = payload.get("meaningful_rows") or []

        previous = self._last_state.get(key, {})
        alerts: List[Tuple[str, str]] = []
        workbook_name = Path(key[0]).name or (key[0] if key[0] else "workbook")
        sheet_label = key[1] if key[1] else "(first sheet)"
        label = f"{workbook_name} — {sheet_label}"

        if status == "Error" and error:
            alerts.append(("error", f"Error detected for {label}: {error}"))
        elif status == "Monitoring started" and previous.get("status") != "Monitoring started":
            alerts.append(("progress", f"Monitoring started for {label}."))
        elif status.startswith("Analyzed"):
            if changed or summary != previous.get("summary"):
                top_summary = summary.splitlines()[0] if isinstance(summary, str) and summary else "Insights refreshed."
                highlight = ""
                if rows:
                    first = rows[0]
                    explanation = first.get("explanation") or first.get("reason")
                    if explanation:
                        highlight = f" Highlight: {explanation}"
                alerts.append(
                    (
                        "insight",
                        f"New insight for {label} (analysis #{analysis_count}). {top_summary}{highlight}",
                    )
                )
            elif analysis_count and analysis_count % 5 == 0 and analysis_count != previous.get("analysis_count"):
                alerts.append(
                    (
                        "progress",
                        f"Progress update: {label} analysed {analysis_count} times with stable telemetry.",
                    )
                )
        elif status == "Stopped" and previous.get("status") != "Stopped":
            alerts.append(("progress", f"Monitoring stopped for {label}."))

        self._last_state[key] = {
            "status": status,
            "summary": summary,
            "analysis_count": analysis_count,
        }

        messages: List[str] = []
        for event_type, text in alerts:
            notify_team(event_type, text)
            entry = f"{timestamp} - {text}"
            self._history.appendleft(entry)
            messages.append(text)
        return messages


class TrendPlot:
    """Render a rolling line-chart of sampled numeric telemetry."""

    def __init__(self, canvas_element: Optional[sg.Canvas]):
        self._canvas_element = canvas_element
        self._figure: Optional[Figure] = None
        self._ax = None
        self._canvas = None
        self.available_columns: List[str] = []
        self.current_column: Optional[str] = None

    def _ready(self) -> bool:
        return _MATPLOTLIB_AVAILABLE and self._canvas_element is not None

    def _ensure_canvas(self) -> bool:
        if not self._ready():
            return False
        if self._figure is None:
            self._figure = Figure(figsize=(5, 3), dpi=100)
            self._ax = self._figure.add_subplot(111)
            self._ax.set_title("Waiting for data")
            self._ax.set_xlabel("Row number")
            self._ax.set_ylabel("Value")
            tk_canvas = self._canvas_element.TKCanvas  # type: ignore[attr-defined]
            self._canvas = FigureCanvasTkAgg(self._figure, master=tk_canvas)
            self._canvas.draw()
            self._canvas.get_tk_widget().pack(side="top", fill="both", expand=1)
        return True

    def update(self, rows: Sequence[dict], preferred_column: Optional[str] = None) -> None:
        if not self._ensure_canvas():
            return

        numeric = self._extract_numeric_columns(rows)
        self.available_columns = sorted(numeric.keys())
        if not self.available_columns:
            self.clear("No numeric telemetry detected yet.")
            return

        if preferred_column and preferred_column in numeric:
            column = preferred_column
        elif self.current_column in numeric:
            column = self.current_column  # type: ignore[assignment]
        else:
            column = self.available_columns[0]
        self.current_column = column

        xs, ys = numeric[column]
        if not xs:
            self.clear("Awaiting numeric data...")
            return

        paired = sorted(zip(xs, ys), key=lambda item: item[0])
        xs_sorted = [item[0] for item in paired]
        ys_sorted = [item[1] for item in paired]

        self._ax.clear()  # type: ignore[union-attr]
        self._ax.set_axis_on()  # type: ignore[union-attr]
        self._ax.plot(xs_sorted, ys_sorted, marker="o", linestyle="-", color="#3A78F2")  # type: ignore[union-attr]
        if len(xs_sorted) > 1:
            self._ax.fill_between(xs_sorted, ys_sorted, alpha=0.12, color="#3A78F2")  # type: ignore[union-attr]
        self._ax.set_title(f"{column} trend (sampled rows)")  # type: ignore[union-attr]
        self._ax.set_xlabel("Row number")  # type: ignore[union-attr]
        self._ax.set_ylabel(column)  # type: ignore[union-attr]
        self._ax.grid(True, linestyle="--", alpha=0.3)  # type: ignore[union-attr]
        self._canvas.draw()  # type: ignore[union-attr]

    def clear(self, message: str = "Select a monitor to visualise.") -> None:
        if not self._ensure_canvas():
            return
        self.available_columns = []
        self.current_column = None
        self._ax.clear()  # type: ignore[union-attr]
        self._ax.text(0.5, 0.5, message, ha="center", va="center", transform=self._ax.transAxes)  # type: ignore[union-attr]
        self._ax.set_axis_off()  # type: ignore[union-attr]
        self._canvas.draw()  # type: ignore[union-attr]

    def _extract_numeric_columns(self, rows: Sequence[dict]) -> Dict[str, Tuple[List[float], List[float]]]:
        columns: Dict[str, Dict[str, List[float]]] = {}
        for entry in rows:
            try:
                row_number = int(entry.get("row_number"))
            except (TypeError, ValueError):
                continue
            values = entry.get("values") or {}
            if not isinstance(values, dict):
                continue
            for column, raw in values.items():
                try:
                    numeric_value = float(raw)
                except (TypeError, ValueError):
                    continue
                bucket = columns.setdefault(column, {"x": [], "y": []})
                bucket["x"].append(row_number)
                bucket["y"].append(numeric_value)

        return {key: (data["x"], data["y"]) for key, data in columns.items() if data["x"]}

# ---------------------------------------------------------------------------
# GUI application
# ---------------------------------------------------------------------------


def _build_table_rows(state: Dict[tuple, dict]) -> List[List[str]]:
    rows: List[List[str]] = []
    for key, payload in sorted(state.items(), key=lambda item: item[0]):
        file_path, sheet = key
        summary = payload.get("summary", "")
        summary_short = summary if len(summary) < 120 else summary[:117] + "..."
        rows.append(
            [
                file_path,
                sheet if sheet else "(first sheet)",
                payload.get("status", ""),
                payload.get("timestamp", ""),
                str(payload.get("analysis_count", 0)),
                payload.get("error", ""),
                summary_short,
            ]
        )
    return rows


def main() -> None:
    sg.theme("SystemDefault")

    headings = [
        "File",
        "Sheet",
        "Status",
        "Last Update",
        "Analyses",
        "Error",
        "AI Summary (truncated)",
    ]

    manager = ExcelMonitorManager(interval=60, sample_rows=25, model="gpt-4o-mini")
    alert_manager = AlertManager()

    settings_row = [
        sg.Text("Polling interval (s)"),
        sg.Input("60", key="-INTERVAL-", size=(6, 1)),
        sg.Text("Sample rows"),
        sg.Input("25", key="-SAMPLE-", size=(6, 1)),
        sg.Text("LLM Model"),
        sg.Input(manager.model, key="-MODEL-", size=(18, 1)),
        sg.Button("Apply", key="-APPLY-"),
    ]

    api_rows = [
        [sg.Text("LLM API URL"), sg.Input(key="-API-URL-", expand_x=True)],
        [sg.Text("LLM API Key"), sg.Input(key="-API-KEY-", password_char="*", expand_x=True)],
    ]

    table_component = sg.Table(
        values=[],
        headings=headings,
        key="-TABLE-",
        expand_x=True,
        expand_y=True,
        auto_size_columns=True,
        justification="left",
        enable_events=True,
        select_mode=sg.TABLE_SELECT_MODE_BROWSE,
    )

    controls_column = [
        [
            sg.Text("Excel file"),
            sg.Input(key="-FILE-", expand_x=True),
            sg.FileBrowse(file_types=(("Excel", "*.xlsx"),)),
            sg.Text("Sheet"),
            sg.Input(key="-SHEET-", size=(14, 1)),
            sg.Button("Add", key="-ADD-"),
        ],
        [table_component],
        [
            sg.Button("Analyze Selected", key="-ANALYZE-"),
            sg.Button("Remove Selected", key="-REMOVE-"),
            sg.Button("Exit"),
        ],
    ]

    detail_frame = sg.Frame(
        "AI Details",
        [
            [
                sg.Multiline(
                    key="-DETAIL-",
                    size=(60, 12),
                    expand_x=True,
                    expand_y=True,
                    autoscroll=True,
                    disabled=True,
                )
            ]
        ],
        expand_x=True,
        expand_y=True,
    )

    if _MATPLOTLIB_AVAILABLE:
        visual_frame = sg.Frame(
            "Trend Visualiser",
            [
                [
                    sg.Text("Numeric column"),
                    sg.Combo(
                        [],
                        key="-COLUMN-SELECT-",
                        readonly=True,
                        enable_events=True,
                        size=(24, 1),
                    ),
                ],
                [
                    sg.Canvas(
                        key="-PLOT-CANVAS-",
                        size=(420, 260),
                        background_color="white",
                    )
                ],
            ],
            expand_x=True,
            expand_y=True,
        )
    else:
        visual_frame = sg.Frame(
            "Trend Visualiser",
            [[sg.Text("Install matplotlib to enable data visualisation.")]],
            expand_x=True,
            expand_y=True,
        )

    alerts_frame = sg.Frame(
        "Smart Alerts",
        [
            [
                sg.Multiline(
                    key="-ALERTS-",
                    size=(60, 8),
                    expand_x=True,
                    expand_y=True,
                    autoscroll=True,
                    disabled=True,
                )
            ]
        ],
        expand_x=True,
        expand_y=True,
    )

    insights_column = [[detail_frame], [visual_frame], [alerts_frame]]

    layout = [
        settings_row,
        *api_rows,
        [
            sg.Column(controls_column, expand_x=True, expand_y=True),
            sg.VSeparator(),
            sg.Column(insights_column, expand_x=True, expand_y=True),
        ],
    ]

    window = sg.Window(
        "Excel AI Monitor",
        layout,
        finalize=True,
        resizable=True,
    )

    canvas_elem = window.get("-PLOT-CANVAS-") if _MATPLOTLIB_AVAILABLE else None
    trend_plot = TrendPlot(canvas_elem)
    if _MATPLOTLIB_AVAILABLE:
        trend_plot.clear()

    table_state: Dict[tuple, dict] = {}
    rows_map: Dict[int, tuple] = {}
    current_key: Optional[tuple] = None
    selected_column: Optional[str] = None
    window["-ALERTS-"].update("")

    def refresh_table() -> None:
        nonlocal rows_map
        table_values = _build_table_rows(table_state)
        rows_map = {idx: key for idx, key in enumerate(sorted(table_state.keys()))}
        window["-TABLE-"].update(values=table_values)
        if current_key and current_key in rows_map.values():
            for idx, key in rows_map.items():
                if key == current_key:
                    window["-TABLE-"].update(select_rows=[idx])
                    break

    def update_detail_for_key(key: Optional[tuple]) -> None:
        nonlocal current_key, selected_column
        previous_key = current_key
        current_key = key
        if not key or key not in table_state:
            window["-DETAIL-"].update("")
            selected_column = None
            window["-TABLE-"].update(select_rows=[])
            if _MATPLOTLIB_AVAILABLE:
                trend_plot.clear()
                window["-COLUMN-SELECT-"].update(values=[], value="")
            return

        payload = table_state.get(key, {})
        detail = {
            "file": key[0],
            "sheet": key[1] if key[1] else "(first sheet)",
            "status": payload.get("status"),
            "timestamp": payload.get("timestamp"),
            "summary": payload.get("summary"),
            "analysis_count": payload.get("analysis_count"),
            "insight_changed": payload.get("insight_changed"),
            "meaningful_rows": payload.get("meaningful_rows"),
            "preview_rows": payload.get("preview_rows"),
            "error": payload.get("error"),
        }
        window["-DETAIL-"].update(_format_rows_for_display(detail))

        if _MATPLOTLIB_AVAILABLE:
            if previous_key != key:
                selected_column = None
            rows = payload.get("preview_rows") or []
            trend_plot.update(rows, selected_column)
            selected_column = trend_plot.current_column
            window["-COLUMN-SELECT-"].update(
                values=trend_plot.available_columns,
                value=trend_plot.current_column or "",
            )

    try:
        while True:
            alerts_triggered = False
            # Drain worker queue
            for message in manager.drain():
                key = (message.get("file", ""), message.get("sheet", ""))
                table_state[key] = message
                if alert_manager.process(message):
                    alerts_triggered = True
                if current_key == key:
                    update_detail_for_key(key)
            refresh_table()
            if alerts_triggered:
                window["-ALERTS-"].update("\n".join(alert_manager.history))

            event, values = window.read(timeout=200)
            if event in (sg.WINDOW_CLOSED, "Exit"):
                break

            if event == "-APPLY-":
                try:
                    interval = int(values["-INTERVAL-"])
                    sample_rows = int(values["-SAMPLE-"])
                except (TypeError, ValueError):
                    sg.popup_error("Interval and sample rows must be integers.", keep_on_top=True)
                    continue

                api_url = values.get("-API-URL-", "")
                api_key = values.get("-API-KEY-", "")
                model = values.get("-MODEL-", manager.model).strip()
                manager.update_global_settings(
                    interval=interval,
                    sample_rows=sample_rows,
                    api_url=api_url,
                    api_key=api_key,
                    model=model,
                )
                sg.popup_no_wait("Settings applied.", keep_on_top=True)

            if event == "-ADD-":
                file_path = values.get("-FILE-", "").strip()
                sheet = values.get("-SHEET-", "").strip()
                if not file_path:
                    sg.popup_error("Please choose an Excel file.", keep_on_top=True)
                    continue
                added = manager.add(file_path, sheet)
                if not added:
                    sg.popup_no_wait("That workbook/sheet is already being monitored.", keep_on_top=True)
                else:
                    sg.popup_no_wait("Workbook added.", keep_on_top=True)

            if event == "-REMOVE-":
                selected_rows = values.get("-TABLE-", [])
                for row in selected_rows:
                    key = rows_map.get(row)
                    if key and manager.remove(key):
                        table_state.pop(key, None)
                        if current_key == key:
                            update_detail_for_key(None)
                refresh_table()

            if event == "-ANALYZE-":
                selected_rows = values.get("-TABLE-", [])
                for row in selected_rows:
                    key = rows_map.get(row)
                    if key:
                        manager.trigger_analysis(key)

            if event == "-TABLE-":
                selection = values.get("-TABLE-", [])
                if selection:
                    row_index = selection[0]
                    update_detail_for_key(rows_map.get(row_index))

            if event == "-COLUMN-SELECT-" and _MATPLOTLIB_AVAILABLE:
                selected_column = values.get("-COLUMN-SELECT-", "") or None
                update_detail_for_key(current_key)

    finally:
        manager.stop_all()
        window.close()


if __name__ == "__main__":  # pragma: no cover - manual execution
    main()

