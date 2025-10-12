# excel_monitor_gui.py
# Parallel Excel Row Growth Monitor (GUI)
# Requirements: pip install PySimpleGUI openpyxl

import threading
import time
from datetime import datetime
from pathlib import Path
from queue import Queue, Empty

import PySimpleGUI as sg
from openpyxl import load_workbook


# ------------------------- Worker (one per file/sheet) ------------------------- #
class TargetWorker(threading.Thread):
    """
    Polls an Excel file/sheet every `interval` seconds, computes row count (ws.max_row),
    and emits status dicts into a shared Queue for the GUI to consume.
    """
    def __init__(self, path, sheet="", interval=10, stall_threshold=2, outq=None):
        super().__init__(daemon=True)
        self.path = str(Path(path))
        self.sheet = sheet  # "" means first sheet
        self.interval = max(1, int(interval))
        self.stall_threshold = max(1, int(stall_threshold))
        self.outq = outq or Queue()
        self._stop = threading.Event()
        self._last = None
        self._stalled = 0

    def stop(self):
        self._stop.set()

    def update_policy(self, interval=None, stall_threshold=None):
        if interval is not None:
            self.interval = max(1, int(interval))
        if stall_threshold is not None:
            self.stall_threshold = max(1, int(stall_threshold))

    @staticmethod
    def _row_count(xlsx_path, sheet_name=""):
        wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        n = ws.max_row or 0
        wb.close()
        return n

    def _emit(self, status, rows="", error=""):
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.outq.put({
            "file": self.path,
            "sheet": self.sheet if self.sheet else "(first sheet)",
            "ts": ts,
            "status": status,
            "rows": rows,
            "error": error
        })

    def run(self):
        # Initial info
        self._emit("Added")

        # Poll loop
        while not self._stop.is_set():
            try:
                if not Path(self.path).exists():
                    raise FileNotFoundError("File not found")

                n = self._row_count(self.path, self.sheet)

                if self._last is None:
                    tag = "Initializing"
                    self._stalled = 0
                else:
                    if n > self._last:
                        tag = f"Growing ({self._last}→{n})"
                        self._stalled = 0
                    elif n == self._last:
                        self._stalled += 1
                        tag = "Stalled" if self._stalled >= self.stall_threshold else "No change"
                    else:
                        tag = f"Decreased ({self._last}→{n})"

                self._last = n
                self._emit(tag, rows=n, error="")

            except Exception as e:
                self._emit("Error", rows="", error=str(e))

            # Sleep in small chunks so stop() is responsive
            total = max(1, self.interval)
            for _ in range(total * 10):
                if self._stop.is_set():
                    break
                time.sleep(0.1)


# ------------------------- Monitor (manages workers) -------------------------- #
class ParallelExcelMonitor:
    """
    Manages many TargetWorker threads. Provides add/remove, global policy updates,
    and a shared output queue for GUI consumption.
    """
    def __init__(self, interval=10, stall_threshold=2):
        self.interval = max(1, int(interval))
        self.stall_threshold = max(1, int(stall_threshold))
        self.outq = Queue()
        self._workers = {}  # key: (path, sheet) -> worker

    @staticmethod
    def _key(path, sheet):
        return (str(Path(path)), sheet or "")

    def add(self, path, sheet=None):
        key = self._key(path, sheet)
        if key in self._workers:
            return False
        w = TargetWorker(
            key[0],
            key[1],
            interval=self.interval,
            stall_threshold=self.stall_threshold,
            outq=self.outq
        )
        self._workers[key] = w
        w.start()
        return True

    def remove(self, key):
        w = self._workers.pop(key, None)
        if w:
            w.stop()
            self.outq.put({
                "file": key[0],
                "sheet": key[1] if key[1] else "(first sheet)",
                "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "status": "Removed",
                "rows": "",
                "error": ""
            })
            return True
        return False

    def list_keys(self):
        return list(self._workers.keys())

    def apply_policy(self, interval, stall_threshold):
        # Update monitor defaults and currently-running workers
        self.interval = max(1, int(interval))
        self.stall_threshold = max(1, int(stall_threshold))
        for w in self._workers.values():
            w.update_policy(self.interval, self.stall_threshold)

    def stop_all(self):
        for w in list(self._workers.values()):
            w.stop()
        self._workers.clear()

    def drain(self):
        items = []
        try:
            while True:
                items.append(self.outq.get_nowait())
        except Empty:
            pass
        return items


# ------------------------------- GUI Front Panel ------------------------------ #
def main():
    sg.theme("SystemDefault")

    headings = ["File", "Sheet", "Rows", "Status", "Last Update", "Note / Error"]

    monitor = ParallelExcelMonitor(interval=10, stall_threshold=2)

    layout = [
        [sg.Text("Monitoring interval (s)"),
         sg.Input("10", size=(6, 1), key="-INTERVAL-"),
         sg.Text("Stall threshold (intervals)"),
         sg.Input("2", size=(6, 1), key="-STALL-"),
         sg.Button("Apply", key="-APPLY-")],

        [sg.Text("Add Excel:"),
         sg.Input(key="-FILE-", expand_x=True),
         sg.FileBrowse(file_types=(("Excel files", "*.xlsx"),)),
         sg.Text("Sheet (optional)"),
         sg.Input(key="-SHEET-", size=(20, 1)),
         sg.Button("Add", key="-ADD-")],

        [sg.Table(values=[],
                  headings=headings,
                  key="-TABLE-",
                  expand_x=True,
                  expand_y=True,
                  auto_size_columns=True,
                  justification="left",
                  enable_events=True,
                  select_mode=sg.TABLE_SELECT_MODE_EXTENDED)],

        [sg.Button("Remove Selected", key="-REMOVE-"),
         sg.Button("Exit")]
    ]

    window = sg.Window(
        "Excel Row Growth Monitor (Parallel)",
        layout,
        finalize=True,
        resizable=True
    )

    # Table backing store and index mapping
    # rows_map: row_index -> key
    rows_map = {}
    table_rows = []

    def rebuild_table_from_state(state_dict):
        """Rebuild table using the latest known state per key (file,sheet)."""
        nonlocal table_rows, rows_map
        table_rows = []
        rows_map = {}
        for i, (key, state) in enumerate(sorted(state_dict.items(), key=lambda x: x[0])):
            path, sheet = key
            table_rows.append([
                path,
                sheet if sheet else "(first sheet)",
                state.get("rows", ""),
                state.get("status", ""),
                state.get("ts", ""),
                state.get("error", ""),
            ])
            rows_map[i] = key
        window["-TABLE-"].update(values=table_rows)

    # We’ll cache the latest status per key for full-table refreshes.
    latest_state = {}

    try:
        while True:
            # Drain worker events and update local state
            for msg in monitor.drain():
                key = (msg["file"], msg["sheet"] if msg["sheet"] != "(first sheet)" else "")
                latest_state[key] = msg

            # Lightweight refresh if there were any changes
            if latest_state:
                rebuild_table_from_state(latest_state)

            event, values = window.read(timeout=200)
            if event in (sg.WINDOW_CLOSED, "Exit"):
                break

            if event == "-ADD-":
                f = values["-FILE-"].strip()
                sheet = values["-SHEET-"].strip()
                if f:
                    added = monitor.add(f, sheet if sheet else None)
                    if not added:
                        sg.popup_no_wait("Already monitoring that file/sheet.", keep_on_top=True)

            if event == "-APPLY-":
                try:
                    new_interval = int(values["-INTERVAL-"])
                    new_stall = int(values["-STALL-"])
                    monitor.apply_policy(new_interval, new_stall)
                    sg.popup_no_wait("Updated polling policy.", keep_on_top=True)
                except Exception:
                    sg.popup_error("Invalid numbers for interval/threshold.", keep_on_top=True)

            if event == "-REMOVE-":
                sel_rows = values["-TABLE-"]
                removed_any = False
                for idx in sorted(sel_rows, reverse=True):
                    key = rows_map.get(idx)
                    if key:
                        if monitor.remove(key):
                            latest_state.pop(key, None)
                            removed_any = True
                if removed_any:
                    rebuild_table_from_state(latest_state)

    finally:
        monitor.stop_all()
        window.close()


if __name__ == "__main__":
    main()
